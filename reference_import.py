"""Idempotent importer for the official hackathon reference workbook."""

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import load_workbook
from sqlmodel import Session, select

from models import AccountingPeriod, Employee, Project, ProjectAssignment, Task


@dataclass
class ImportSummary:
    employees_created: int = 0
    employees_updated: int = 0
    projects_created: int = 0
    projects_updated: int = 0
    tasks_created: int = 0
    tasks_updated: int = 0
    periods_created: int = 0
    periods_updated: int = 0
    assignments_created: int = 0
    warnings: list[str] = field(default_factory=list)

    @property
    def changed_records(self) -> int:
        return sum(
            (
                self.employees_created,
                self.employees_updated,
                self.projects_created,
                self.projects_updated,
                self.tasks_created,
                self.tasks_updated,
                self.periods_created,
                self.periods_updated,
                self.assignments_created,
            )
        )


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _expense_type(value: object) -> str:
    return "CapEx" if "capex" in _text(value).casefold() else "OpEx"


def _identifier(prefix: str, value: str) -> str:
    slug = re.sub(r"[^A-Z0-9]+", "-", value.upper()).strip("-")
    return f"{prefix}-{slug}"[:50]


def _date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(_text(value))


def import_reference_workbook(session: Session, workbook_path: str | Path) -> ImportSummary:
    """Upsert people, projects, taxonomy, classification, and fiscal periods."""
    path = Path(workbook_path)
    if not path.is_file():
        raise ValueError(f"Workbook not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    required = {
        "Partner Directory",
        "Project List",
        "Categories",
        "CapEx - OpEx Classification",
        "FY26 -FY27 Calendar",
    }
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ValueError("Workbook is missing sheets: " + ", ".join(sorted(missing)))

    summary = ImportSummary()
    _import_employees(session, workbook["Partner Directory"], summary)
    classifications = _read_classifications(
        workbook["CapEx - OpEx Classification"]
    )
    project_tasks, non_project_tasks = _read_categories(workbook["Categories"])
    _import_projects(
        session,
        workbook["Project List"],
        project_tasks,
        non_project_tasks,
        classifications,
        summary,
    )
    _import_periods(session, workbook["FY26 -FY27 Calendar"], summary)
    _seed_demo_assignments(session, summary)
    session.commit()
    return summary


def _import_employees(session: Session, sheet, summary: ImportSummary) -> None:
    existing = {
        employee.partner_id: employee
        for employee in session.exec(select(Employee)).all()
        if employee.partner_id
    }
    seen: dict[str, tuple[str, str]] = {}
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        name, role, team, partner_id, email, employment, status = row[:7]
        partner_id = _text(partner_id)
        if not partner_id:
            continue
        signature = (_text(name), _text(team))
        if partner_id in seen:
            if seen[partner_id] != signature:
                summary.warnings.append(
                    f"Partner Directory row {row_number}: duplicate {partner_id} "
                    f"conflicts with an earlier row; the first record was kept."
                )
            else:
                summary.warnings.append(
                    f"Partner Directory row {row_number}: exact duplicate "
                    f"{partner_id} was skipped."
                )
            continue
        seen[partner_id] = signature
        employee = existing.get(partner_id)
        created = employee is None
        if employee is None:
            employee = Employee(partner_id=partner_id)
            session.add(employee)
            existing[partner_id] = employee
        employee.name = _text(name)[:100]
        employee.job_title = _text(role)[:100]
        employee.department = _text(team)[:100]
        employee.email = _text(email)[:254]
        employee.engagement_type = (
            "Contractor" if _text(employment).casefold() == "contractor" else "Full-time"
        )
        employee.active = _text(status).casefold() == "active"
        employee.expected_weekly_hours = 40
        summary.employees_created += int(created)
        summary.employees_updated += int(not created)


def _read_classifications(sheet) -> dict[str, str]:
    classifications: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=4, max_row=14, values_only=True):
        code = _text(row[0])
        if code:
            classifications[code.casefold()] = _expense_type(row[1])
    return classifications


def _read_categories(sheet) -> tuple[list[tuple[str, str]], list[tuple[str, str, str]]]:
    project_tasks = [
        (_text(row[1]), _text(row[2]))
        for row in sheet.iter_rows(min_row=4, max_row=14, values_only=True)
        if _text(row[1])
    ]
    non_project_tasks = [
        (_text(row[0]), _text(row[1]), _text(row[2]))
        for row in sheet.iter_rows(min_row=20, max_row=28, values_only=True)
        if _text(row[1])
    ]
    return project_tasks, non_project_tasks


def _upsert_tasks(
    session: Session,
    project: Project,
    definitions: list[tuple[str, str]],
    classifications: dict[str, str],
    summary: ImportSummary,
) -> None:
    session.flush()
    existing = {
        task.name.casefold(): task
        for task in session.exec(select(Task).where(Task.project_id == project.id)).all()
    }
    for name, description in definitions:
        task = existing.get(name.casefold())
        created = task is None
        if task is None:
            task = Task(project_id=project.id, name=name)
            session.add(task)
            existing[name.casefold()] = task
        task.description = description[:500]
        task.expense_type = (
            "OpEx"
            if project.work_type == "Non-project"
            else classifications.get(name.casefold(), "OpEx")
        )
        task.active = True
        summary.tasks_created += int(created)
        summary.tasks_updated += int(not created)


def _import_projects(
    session: Session,
    sheet,
    project_tasks: list[tuple[str, str]],
    non_project_tasks: list[tuple[str, str, str]],
    classifications: dict[str, str],
    summary: ImportSummary,
) -> None:
    existing = {
        project.identifier: project
        for project in session.exec(select(Project)).all()
        if project.identifier
    }
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, number, status, notes, manager = row[:5]
        identifier = _text(number)
        if not identifier or not _text(name):
            continue
        project = existing.get(identifier)
        created = project is None
        if project is None:
            project = Project(identifier=identifier)
            session.add(project)
            existing[identifier] = project
        project.name = _text(name)[:100]
        project.status = _text(status)[:20] or "Active"
        project.active = project.status.casefold() == "active"
        project.notes = _text(notes)[:500]
        project.project_manager = _text(manager)[:100]
        project.work_type = "Project"
        project.project_type = "Hackathon reference"
        project.cost_type = "Mixed"
        summary.projects_created += int(created)
        summary.projects_updated += int(not created)
        _upsert_tasks(session, project, project_tasks, classifications, summary)

    grouped: dict[str, list[tuple[str, str]]] = {}
    for group, code, description in non_project_tasks:
        grouped.setdefault(group, []).append((code, description))
    for group, definitions in grouped.items():
        identifier = _identifier("NP", group)
        project = existing.get(identifier)
        created = project is None
        if project is None:
            project = Project(identifier=identifier)
            session.add(project)
            existing[identifier] = project
        project.name = group[:100]
        project.client = "Internal"
        project.work_type = "Non-project"
        project.project_type = "Non-project category"
        project.status = "Active"
        project.active = True
        project.cost_type = "Operating"
        summary.projects_created += int(created)
        summary.projects_updated += int(not created)
        _upsert_tasks(session, project, definitions, classifications, summary)


def _import_periods(session: Session, sheet, summary: ImportSummary) -> None:
    existing = {
        (period.start_date, period.end_date): period
        for period in session.exec(select(AccountingPeriod)).all()
    }
    for fiscal_year, offset in (("FY26", 0), ("FY27", 8)):
        for row in sheet.iter_rows(min_row=5, max_row=16, values_only=True):
            quarter, number, start, end, weeks = row[offset : offset + 5]
            start_date, end_date = _date(start), _date(end)
            period = existing.get((start_date, end_date))
            created = period is None
            if period is None:
                period = AccountingPeriod(
                    name=f"{fiscal_year} period {int(number):02d}",
                    start_date=start_date,
                    end_date=end_date,
                )
                session.add(period)
                existing[(start_date, end_date)] = period
            period.fiscal_year = fiscal_year
            period.quarter = int(quarter)
            period.period_number = int(number)
            period.week_count = int(weeks)
            summary.periods_created += int(created)
            summary.periods_updated += int(not created)


def _seed_demo_assignments(session: Session, summary: ImportSummary) -> None:
    """Create a repeatable demo allocation because the workbook has no assignment tab."""
    employees = list(
        session.exec(
            select(Employee).where(Employee.partner_id != "", Employee.active.is_(True))
        ).all()
    )
    projects = list(
        session.exec(
            select(Project)
            .where(Project.work_type == "Project", Project.active.is_(True))
            .order_by(Project.identifier)
        ).all()
    )
    existing = {
        (item.employee_id, item.project_id)
        for item in session.exec(select(ProjectAssignment)).all()
    }
    for employee in employees:
        digits = int("".join(character for character in employee.partner_id if character.isdigit()) or 0)
        selected = {
            projects[(digits + offset * 5) % len(projects)].id
            for offset in range(min(4, len(projects)))
        }
        if employee.partner_id == "EMP0078":
            selected.update(project.id for project in projects if project.identifier == "11")
        if employee.partner_id == "EMP0059":
            selected.difference_update(
                project.id for project in projects if project.identifier == "24"
            )
        for project_id in selected:
            if (employee.id, project_id) in existing:
                continue
            session.add(
                ProjectAssignment(
                    employee_id=employee.id,
                    project_id=project_id,
                    assigned_by="Workbook demo allocation",
                )
            )
            existing.add((employee.id, project_id))
            summary.assignments_created += 1
